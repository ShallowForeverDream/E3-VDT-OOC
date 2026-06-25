from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


FIELD_COLUMNS = {
    "gold_entity": "entity",
    "gold_location": "location",
    "gold_time": "time",
    "gold_event_type": "event_type",
    "gold_relation": "relation",
    "gold_global": "different_event",
    "gold_context_omission": "context_omission",
    "gold_evidence_insufficient": "evidence_insufficient",
}


def truthy(x: Any) -> bool:
    return str(x or "").strip().upper() in {"Y", "YES", "TRUE", "1", "是"}


def main() -> None:
    ap = argparse.ArgumentParser(description="Import the XLSX annotation workbook back to attribution_eval_set.jsonl.")
    ap.add_argument("--xlsx", default="examples/attribution_eval_candidates_annotate.xlsx")
    ap.add_argument("--output", default="examples/attribution_eval_set.jsonl")
    ap.add_argument("--sheet", default="Annotation")
    ap.add_argument("--done-only", action="store_true", help="Export only rows with annotation_status=done")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb[args.sheet]
    headers = [str(c.value or "").strip() for c in ws[1]]
    idx = {name: i + 1 for i, name in enumerate(headers)}
    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        sample_id = ws.cell(r, idx["sample_id"]).value if "sample_id" in idx else None
        if not sample_id:
            continue
        status = str(ws.cell(r, idx.get("annotation_status", 0)).value or "").strip()
        if args.done_only and status != "done":
            continue
        fields = []
        for col_name, field in FIELD_COLUMNS.items():
            if col_name in idx and truthy(ws.cell(r, idx[col_name]).value):
                fields.append(field)
        row: Dict[str, Any] = {
            "sample_id": str(sample_id),
            "image_id": str(ws.cell(r, idx["image_id"]).value or "") if "image_id" in idx else "",
            "text_id": str(ws.cell(r, idx["text_id"]).value or "") if "text_id" in idx else "",
            "split": ws.cell(r, idx["split"]).value if "split" in idx else "",
            "domain": ws.cell(r, idx["domain"]).value if "domain" in idx else "",
            "label": 1 if str(ws.cell(r, idx["label"]).value or "") == "OOC" else 0 if str(ws.cell(r, idx["label"]).value or "") == "Non-OOC" else None,
            "current_caption": ws.cell(r, idx["current_caption"]).value if "current_caption" in idx else "",
            "true_image_context": ws.cell(r, idx["true_image_context"]).value if "true_image_context" in idx else "",
            "gold_mismatch_type": str(ws.cell(r, idx["gold_mismatch_type"]).value or "").strip() if "gold_mismatch_type" in idx else "",
            "gold_conflict_fields": fields,
            "annotator": str(ws.cell(r, idx["annotator"]).value or "").strip() if "annotator" in idx else "",
            "rationale": str(ws.cell(r, idx["rationale"]).value or "").strip() if "rationale" in idx else "",
            "annotation_status": status or "todo",
        }
        rows.append(row)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    print(json.dumps({"output": str(out), "records": len(rows), "done_only": args.done_only}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
