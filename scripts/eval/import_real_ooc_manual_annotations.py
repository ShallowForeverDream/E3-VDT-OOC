from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple


TYPE_ZH_TO_EN = {
    "主体/人物错配": "entity mismatch",
    "主体人物错配": "entity mismatch",
    "人物错配": "entity mismatch",
    "地点错配": "location mismatch",
    "时间错配": "temporal mismatch",
    "事件类型错配": "event-type mismatch",
    "关系/动作错配": "relation mismatch",
    "关系动作错配": "relation mismatch",
    "完全不同事件": "different-event mismatch",
    "上下文遗漏": "context omission",
    "证据不足": "uncertain / evidence insufficient",
    "证据不足/不确定": "uncertain / evidence insufficient",
    "不确定": "uncertain / evidence insufficient",
    "无明显语义冲突/泛化配图": "benign illustrative image",
    "泛化配图": "benign illustrative image",
}

FIELD_COLUMNS = [
    ("主体/人物冲突", "entity"),
    ("地点冲突", "location"),
    ("时间冲突", "time"),
    ("事件类型冲突", "event_type"),
    ("关系/动作冲突", "relation"),
    ("上下文遗漏", "context_omission"),
    ("证据不足", "evidence_insufficient"),
]

DEFAULT_FIELDS_BY_TYPE = {
    "entity mismatch": ["entity"],
    "location mismatch": ["location"],
    "temporal mismatch": ["time"],
    "event-type mismatch": ["event_type"],
    "relation mismatch": ["relation"],
    "different-event mismatch": ["entity", "event_type", "relation"],
    "context omission": ["context_omission"],
    "uncertain / evidence insufficient": ["evidence_insufficient"],
    "benign illustrative image": [],
}

STATUS_ZH_TO_EN = {
    "已完成": "done",
    "待标注": "todo",
    "不确定": "uncertain",
    "跳过": "skip",
}


def clean(x: Any) -> str:
    return str(x if x is not None else "").strip()


def yes(x: Any) -> bool:
    return clean(x).lower() in {"是", "y", "yes", "true", "1"}


def norm_type_zh(x: Any) -> str:
    s = re.sub(r"\s+", "", clean(x))
    return TYPE_ZH_TO_EN.get(s, clean(x))


def read_sheet_rows(path: Path, preferred_prefix: str = "中文标注") -> Tuple[str, List[Dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise SystemExit("缺少 openpyxl。请运行：python -m pip install openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.startswith(preferred_prefix)), wb.sheetnames[0])
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return sheet_name, []
    headers = [clean(h) for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for raw in rows[1:]:
        row = {h: raw[i] if i < len(raw) else "" for i, h in enumerate(headers) if h}
        if clean(row.get("样本ID")):
            out.append(row)
    return sheet_name, out


def read_backup_rows(path: Path) -> Dict[str, Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise SystemExit("缺少 openpyxl。请运行：python -m pip install openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = next(
        (s for s in wb.sheetnames if "英文" in s or "备份" in s),
        "",
    )
    if not sheet_name:
        return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [clean(h) for h in rows[0]]
    out: Dict[str, Dict[str, Any]] = {}
    for raw in rows[1:]:
        row = {h: raw[i] if i < len(raw) else "" for i, h in enumerate(headers) if h}
        sid = clean(row.get("sample_id") or row.get("样本ID"))
        if sid:
            out[sid] = row
    return out


def backup_value(backup: Dict[str, Any], *names: str) -> str:
    for n in names:
        if clean(backup.get(n)):
            return clean(backup.get(n))
    return ""


def rationale_warning(rationale: str) -> str:
    s = clean(rationale)
    if not s:
        return "missing_rationale"
    if re.fullmatch(r"\d+(?:\.0)?", s):
        return "numeric_rationale"
    return ""


def canonicalize_row(row: Dict[str, Any], backup: Dict[str, Any], source_path: Path, batch: str, idx: int) -> Dict[str, Any]:
    sid = clean(row.get("样本ID"))
    typ_zh = clean(row.get("人工主错配类型"))
    typ_en = norm_type_zh(typ_zh)
    fields = [field for col, field in FIELD_COLUMNS if yes(row.get(col))]
    if not fields:
        fields = list(DEFAULT_FIELDS_BY_TYPE.get(typ_en, []))
    rationale = clean(row.get("人工判断依据（中文）"))
    status_zh = clean(row.get("标注状态"))
    status = STATUS_ZH_TO_EN.get(status_zh, status_zh or "done")
    current_en = backup_value(backup, "current_caption_en", "英文待检测文本")
    true_en = backup_value(backup, "true_image_context_en", "英文图片真实上下文")
    out = {
        "sample_id": sid,
        "batch": batch,
        "idx_in_batch": idx,
        "data_label": clean(row.get("数据标签")) or "真实OOC",
        "label": 1,
        "vdt_label": "OOC",
        "vdt_score": 0.9,
        "split": clean(row.get("数据划分")),
        "domain": clean(row.get("新闻域")),
        "current_caption": current_en or clean(row.get("待检测文本（中文）")),
        "true_image_context": true_en or clean(row.get("图片真实上下文（中文）")),
        "current_caption_zh": clean(row.get("待检测文本（中文）")),
        "true_image_context_zh": clean(row.get("图片真实上下文（中文）")),
        "gold_mismatch_type": typ_en,
        "gold_mismatch_type_zh": typ_zh,
        "gold_conflict_fields": fields,
        "annotation_status": status,
        "annotation_status_zh": status_zh,
        "annotator": clean(row.get("标注人")),
        "rationale": rationale,
        "rationale_zh": rationale,
        "rationale_warning": rationale_warning(rationale),
        "image_id": clean(row.get("图片ID")),
        "text_id": clean(row.get("文本ID")) or sid,
        "image_path": clean(row.get("图片路径")),
        "image_topic": clean(row.get("图片主题")),
        "negative_source": clean(row.get("负样本来源")),
        "image_source": clean(row.get("图片来源")),
        "model_reference_mismatch_type_zh": clean(row.get("模型参考错配类型") or backup.get("模型参考错配类型")),
        "model_reference_conflict_fields_zh": clean(row.get("模型参考冲突字段") or backup.get("模型参考冲突字段")),
        "newsclippings_file": backup_value(backup, "newsclippings_file"),
        "source_xlsx": str(source_path),
        "source_sheet_batch": batch,
        "notes": clean(row.get("备注")),
    }
    return out


def write_jsonl(path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: json.dumps(v, ensure_ascii=False) if isinstance(v, list) else v for k, v in row.items()})


def summarize(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    ids = [r["sample_id"] for r in rows]
    duplicate_ids = sorted([sid for sid, n in Counter(ids).items() if n > 1])
    type_counts = Counter(r.get("gold_mismatch_type") for r in rows)
    field_counts: Counter[str] = Counter()
    for r in rows:
        field_counts.update(r.get("gold_conflict_fields") or [])
    domain_counts = Counter(r.get("domain") for r in rows)
    batch_counts = Counter(r.get("batch") for r in rows)
    warning_counts = Counter(r.get("rationale_warning") or "ok" for r in rows)
    completed = sum(1 for r in rows if r.get("annotation_status") == "done")
    return {
        "records": len(rows),
        "completed": completed,
        "duplicate_sample_ids": duplicate_ids,
        "type_distribution": dict(type_counts.most_common()),
        "field_distribution": dict(field_counts.most_common()),
        "domain_distribution": dict(domain_counts.most_common()),
        "batch_distribution": dict(batch_counts.most_common()),
        "rationale_warning_counts": dict(warning_counts.most_common()),
        "dominant_type": type_counts.most_common(1)[0][0] if type_counts else "",
        "dominant_type_ratio": (type_counts.most_common(1)[0][1] / len(rows)) if rows and type_counts else 0.0,
    }


def main() -> None:
    ap = argparse.ArgumentParser(description="Import two Chinese real-OOC manual annotation workbooks into canonical JSONL/CSV.")
    ap.add_argument(
        "--xlsx",
        nargs="+",
        default=[
            "outputs/ooc50_annotation_20260626/真实OOC人工标注50条_中文.xlsx",
            "outputs/ooc50_annotation_batch2_20260626/真实OOC人工标注第2批50条_中文.xlsx",
        ],
    )
    ap.add_argument("--output", default="examples/real_ooc_attribution_eval_set.jsonl")
    ap.add_argument("--csv", default="examples/real_ooc_manual_100_canonical.csv")
    ap.add_argument("--summary", default="outputs/real_ooc_manual_label_stats.json")
    ap.add_argument("--summary-public", default="examples/real_ooc_manual_100_summary.json")
    args = ap.parse_args()

    rows: List[Dict[str, Any]] = []
    for batch_idx, raw_path in enumerate(args.xlsx, start=1):
        path = Path(raw_path)
        if not path.exists():
            raise FileNotFoundError(path)
        _, main_rows = read_sheet_rows(path)
        backups = read_backup_rows(path)
        batch = "batch1" if batch_idx == 1 else f"batch{batch_idx}"
        for idx, row in enumerate(main_rows, start=1):
            sid = clean(row.get("样本ID"))
            rows.append(canonicalize_row(row, backups.get(sid, {}), path, batch, idx))

    summary = summarize(rows)
    write_jsonl(Path(args.output), rows)
    write_csv(Path(args.csv), rows)
    out = Path(args.summary)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    public = Path(args.summary_public)
    if str(public):
        public.parent.mkdir(parents=True, exist_ok=True)
        public.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    csv_summary = out.with_suffix(".csv")
    with csv_summary.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["section", "key", "value"])
        writer.writerow(["summary", "records", summary["records"]])
        writer.writerow(["summary", "completed", summary["completed"]])
        writer.writerow(["summary", "dominant_type", summary["dominant_type"]])
        writer.writerow(["summary", "dominant_type_ratio", summary["dominant_type_ratio"]])
        for section in ["type_distribution", "field_distribution", "domain_distribution", "batch_distribution", "rationale_warning_counts"]:
            for k, v in summary[section].items():
                writer.writerow([section, k, v])
    print(json.dumps({"output": args.output, "csv": args.csv, "summary": args.summary, **summary}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
