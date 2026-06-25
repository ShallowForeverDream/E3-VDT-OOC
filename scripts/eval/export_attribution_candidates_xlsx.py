from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


MISMATCH_TYPES = [
    "benign illustrative image",
    "entity mismatch",
    "location mismatch",
    "temporal mismatch",
    "event-type mismatch",
    "relation mismatch",
    "different-event mismatch",
    "context omission",
    "uncertain / evidence insufficient",
]

STATUS_VALUES = ["todo", "done", "uncertain", "skip"]
YN_VALUES = ["N", "Y"]
ANNOTATORS = ["组长", "复现负责人", "系统负责人", "报告负责人"]
CONFLICT_FIELDS = ["entity", "location", "time", "event_type", "relation", "context_omission", "evidence_insufficient"]


def read_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            if line.strip():
                rows.append(json.loads(line))
    return rows


def index_by_sample(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    return {str(r.get("sample_id") or r.get("id") or r.get("image_id") or ""): r for r in rows}


def as_csv(xs: Any) -> str:
    if not xs:
        return ""
    if isinstance(xs, str):
        return xs
    return ", ".join(str(x) for x in xs if str(x).strip())


def yn_for(fields: Any, field: str) -> str:
    if isinstance(fields, str):
        parts = {x.strip() for x in fields.split(",")}
    else:
        parts = {str(x).strip() for x in (fields or [])}
    return "Y" if field in parts else "N"


def nli_summary(row: Dict[str, Any]) -> str:
    field_nli = row.get("field_nli") or {}
    parts = []
    for field in ["entity", "location", "time", "event_type", "relation"]:
        item = field_nli.get(field, {})
        if isinstance(item, dict):
            label = item.get("label", "")
            scores = item.get("scores", {}) if isinstance(item.get("scores", {}), dict) else {}
            c = scores.get("contradiction")
            if isinstance(c, (int, float)):
                parts.append(f"{field}:{label}({c:.2f})")
            else:
                parts.append(f"{field}:{label}")
    return "; ".join(parts)


def build_workbook(candidate_rows: List[Dict[str, Any]], pred_rows: List[Dict[str, Any]], output: Path) -> None:
    pred_by_id = index_by_sample(pred_rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "Annotation"
    options = wb.create_sheet("Options")
    readme = wb.create_sheet("README")

    headers = [
        "idx",
        "sample_id",
        "label",
        "split",
        "domain",
        "current_caption",
        "true_image_context",
        "v2_mismatch_type",
        "v2_conflict_fields",
        "weak_mismatch_type",
        "weak_conflict_fields",
        "gold_mismatch_type",
        "gold_entity",
        "gold_location",
        "gold_time",
        "gold_event_type",
        "gold_relation",
        "gold_context_omission",
        "gold_evidence_insufficient",
        "annotation_status",
        "annotator",
        "rationale",
        "nli_summary",
        "evidence_relevance",
        "image_id",
        "text_id",
        "generator",
    ]
    ws.append(headers)

    for idx, row in enumerate(candidate_rows, start=1):
        sid = str(row.get("sample_id", ""))
        pred = pred_by_id.get(sid, row)
        v2_fields = pred.get("v2_conflict_fields", [])
        weak_fields = row.get("weak_conflict_fields", pred.get("weak_conflict_fields", []))
        evidence = pred.get("evidence_relevance", {})
        evidence_score = evidence.get("evidence_relevance", "") if isinstance(evidence, dict) else ""
        label = row.get("label", "")
        label_name = "OOC" if label == 1 else "Non-OOC" if label == 0 else str(label)
        ws.append([
            idx,
            sid,
            label_name,
            row.get("split", ""),
            row.get("domain", ""),
            row.get("current_caption", ""),
            row.get("true_image_context", ""),
            pred.get("v2_mismatch_type", ""),
            as_csv(v2_fields),
            row.get("weak_mismatch_type", pred.get("weak_mismatch_type", "")),
            as_csv(weak_fields),
            row.get("gold_mismatch_type", ""),
            yn_for(row.get("gold_conflict_fields", []), "entity"),
            yn_for(row.get("gold_conflict_fields", []), "location"),
            yn_for(row.get("gold_conflict_fields", []), "time"),
            yn_for(row.get("gold_conflict_fields", []), "event_type"),
            yn_for(row.get("gold_conflict_fields", []), "relation"),
            yn_for(row.get("gold_conflict_fields", []), "context_omission"),
            yn_for(row.get("gold_conflict_fields", []), "evidence_insufficient"),
            row.get("annotation_status", "todo") or "todo",
            row.get("annotator", ""),
            row.get("rationale", ""),
            nli_summary(pred),
            evidence_score,
            row.get("image_id", pred.get("image_id", "")),
            row.get("text_id", pred.get("text_id", "")),
            row.get("generator", pred.get("generator", "")),
        ])

    # Options sheet.
    options["A1"] = "mismatch_type_options"
    for i, val in enumerate(MISMATCH_TYPES, start=2):
        options[f"A{i}"] = val
    options["C1"] = "Y/N"
    for i, val in enumerate(YN_VALUES, start=2):
        options[f"C{i}"] = val
    options["E1"] = "annotation_status"
    for i, val in enumerate(STATUS_VALUES, start=2):
        options[f"E{i}"] = val
    options["G1"] = "annotator"
    for i, val in enumerate(ANNOTATORS, start=2):
        options[f"G{i}"] = val
    options["I1"] = "conflict_fields"
    for i, val in enumerate(CONFLICT_FIELDS, start=2):
        options[f"I{i}"] = val

    # README sheet.
    readme_rows = [
        ["VDT-COVE-Attr 人工归因标注表"],
        [""],
        ["填写规则"],
        ["1. gold_mismatch_type：从下拉框选择人工认为最主要的错配类型。"],
        ["2. gold_entity / gold_location / ...：每个冲突字段用下拉框选择 Y/N。Excel 原生不支持稳定多选下拉，所以用多个 Y/N 字段替代。"],
        ["3. annotation_status：标完选择 done；拿不准选 uncertain；不适合评估选 skip。"],
        ["4. rationale：简短写人工判断依据，方便答辩追问时解释。"],
        [""],
        ["推荐标注口径"],
        ["Non-OOC：gold_mismatch_type 选 benign illustrative image，所有 gold_* 字段选 N。"],
        ["OOC 且证据不足：gold_mismatch_type 选 uncertain / evidence insufficient，gold_evidence_insufficient 选 Y。"],
        ["OOC 且能定位字段：选择对应 mismatch_type，并把相关 gold_* 字段选 Y。"],
        ["OOC 且整件事都对不上：gold_mismatch_type 选 different-event mismatch，并把多个冲突字段选 Y。"],
        [""],
        ["标完后导回 JSONL："],
        ["python scripts\\eval\\import_attribution_xlsx.py --xlsx examples\\attribution_eval_candidates_annotate.xlsx --output examples\\attribution_eval_set.jsonl"],
    ]
    for row in readme_rows:
        readme.append(row)

    # Styling.
    dark = "1F2937"
    blue = "2563EB"
    light_blue = "DBEAFE"
    green = "DCFCE7"
    yellow = "FEF3C7"
    red = "FEE2E2"
    gray = "F3F4F6"
    border_color = "D1D5DB"
    thin = Side(style="thin", color=border_color)
    for sheet in [ws, options, readme]:
        sheet.sheet_view.showGridLines = False

    header = ws[1]
    for cell in header:
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=blue))

    widths = {
        "A": 6, "B": 13, "C": 10, "D": 9, "E": 14,
        "F": 44, "G": 44, "H": 22, "I": 24, "J": 22, "K": 24,
        "L": 28, "M": 12, "N": 13, "O": 10, "P": 14, "Q": 12, "R": 20, "S": 24,
        "T": 18, "U": 16, "V": 42, "W": 46, "X": 16, "Y": 14, "Z": 14, "AA": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
        row[11].fill = PatternFill("solid", fgColor=yellow)  # gold_mismatch_type
        for col_idx in range(12, 19):
            row[col_idx].fill = PatternFill("solid", fgColor=yellow)
        row[19].fill = PatternFill("solid", fgColor=yellow)
        row[20].fill = PatternFill("solid", fgColor=yellow)
        row[21].fill = PatternFill("solid", fgColor=yellow)

    # Color label rows lightly by OOC/Non-OOC.
    ws.conditional_formatting.add(f"C2:C{ws.max_row}", FormulaRule(formula=['$C2="OOC"'], fill=PatternFill("solid", fgColor=red)))
    ws.conditional_formatting.add(f"C2:C{ws.max_row}", FormulaRule(formula=['$C2="Non-OOC"'], fill=PatternFill("solid", fgColor=green)))
    ws.conditional_formatting.add(f"T2:T{ws.max_row}", FormulaRule(formula=['$T2="done"'], fill=PatternFill("solid", fgColor=green)))
    ws.conditional_formatting.add(f"T2:T{ws.max_row}", FormulaRule(formula=['$T2="uncertain"'], fill=PatternFill("solid", fgColor=yellow)))

    ws.freeze_panes = "F2"
    ws.auto_filter.ref = f"A1:AA{ws.max_row}"
    ws.row_dimensions[1].height = 34
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 92

    # Data validation dropdowns.
    max_row = max(ws.max_row, 2)
    dv_type = DataValidation(type="list", formula1=f"=Options!$A$2:$A${len(MISMATCH_TYPES)+1}", allow_blank=True)
    dv_yn = DataValidation(type="list", formula1=f"=Options!$C$2:$C${len(YN_VALUES)+1}", allow_blank=False)
    dv_status = DataValidation(type="list", formula1=f"=Options!$E$2:$E${len(STATUS_VALUES)+1}", allow_blank=False)
    dv_annotator = DataValidation(type="list", formula1=f"=Options!$G$2:$G${len(ANNOTATORS)+1}", allow_blank=True)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_yn)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_annotator)
    dv_type.add(f"L2:L{max_row}")
    dv_yn.add(f"M2:S{max_row}")
    dv_status.add(f"T2:T{max_row}")
    dv_annotator.add(f"U2:U{max_row}")

    # Excel table.
    tab = Table(displayName="AttributionAnnotationTable", ref=f"A1:AA{ws.max_row}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Comments for editable columns.
    ws["L1"].comment = Comment("人工选择主错配类型。", "Codex")
    ws["M1"].comment = Comment("字段是否存在冲突：Y/N。多字段冲突可以多个列选 Y。", "Codex")
    ws["T1"].comment = Comment("标注完成后选 done；不确定选 uncertain。", "Codex")
    ws["V1"].comment = Comment("简短写判断依据，便于报告/答辩解释。", "Codex")

    # Options/readme formatting.
    for sheet in [options, readme]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.column_dimensions["A"].width = 56
        sheet.column_dimensions["C"].width = 16
        sheet.column_dimensions["E"].width = 20
        sheet.column_dimensions["G"].width = 18
        sheet.column_dimensions["I"].width = 24
        sheet["A1"].font = Font(bold=True, size=14, color=dark)
    for cell in options[1]:
        cell.fill = PatternFill("solid", fgColor=gray)
        cell.font = Font(bold=True)
    readme["A1"].font = Font(bold=True, size=16, color=blue)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    ap = argparse.ArgumentParser(description="Export attribution JSONL candidates to an XLSX annotation workbook with dropdowns.")
    ap.add_argument("--candidates", default="examples/attribution_eval_candidates.jsonl")
    ap.add_argument("--pred", default="outputs/field_nli_attribution_v2.jsonl")
    ap.add_argument("--output", default="examples/attribution_eval_candidates_annotate.xlsx")
    args = ap.parse_args()

    candidates = read_jsonl(Path(args.candidates))
    pred_rows = read_jsonl(Path(args.pred)) if Path(args.pred).exists() else []
    build_workbook(candidates, pred_rows, Path(args.output))
    print(json.dumps({"output": args.output, "records": len(candidates)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
