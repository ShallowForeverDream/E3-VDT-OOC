from __future__ import annotations

import argparse
import json
import random
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

LABELS = [
    "location mismatch",
    "temporal mismatch",
    "entity mismatch",
    "different-event mismatch",
    "event-type mismatch",
    "relation mismatch",
    "global/uncontrolled mismatch",
    "evidence insufficient",
    "benign illustrative image",
]
YN = ["Y", "N"]
STATUS = ["todo", "done", "uncertain", "skip"]
ANNOTATORS = ["组长", "复现负责人", "系统负责人", "报告负责人"]


def read_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not path.exists():
        return rows
    with path.open(encoding="utf-8") as f:
        for line in f:
            if line.strip():
                rows.append(json.loads(line))
    return rows


def write_jsonl(path: Path, rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def key_of(row: Dict[str, Any]) -> str:
    return str(row.get("sample_id") or row.get("id") or row.get("image_id") or "").strip()


def is_ooc(row: Dict[str, Any]) -> bool:
    label = row.get("label")
    if label == 1:
        return True
    if isinstance(label, bool):
        return bool(label)
    if isinstance(label, str):
        return label.strip().lower() in {"1", "true", "ooc", "falsified", "mismatch"}
    return False


def info_score(row: Dict[str, Any], pred: Optional[Dict[str, Any]]) -> float:
    cur = str(row.get("current_caption") or "")
    tru = str(row.get("true_image_context") or "")
    score = min(len(cur), 200) + min(len(tru), 200)
    if pred:
        ev = pred.get("evidence_relevance") or {}
        score += float(ev.get("evidence_relevance") or 0) * 100
        score += 20 * len(pred.get("v2_conflict_fields") or [])
    return score


def export_xlsx(path: Path, rows: List[Dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as exc:  # pragma: no cover
        print(f"openpyxl unavailable, skip xlsx export: {exc}")
        return

    headers = [
        "sample_id",
        "image_id",
        "domain",
        "current_caption",
        "true_image_context",
        "pred_mismatch_type",
        "pred_conflict_fields",
        "gold_mismatch_type",
        "gold_entity",
        "gold_location",
        "gold_time",
        "gold_event_type",
        "gold_relation",
        "gold_global",
        "gold_evidence_insufficient",
        "rationale",
        "annotation_status",
        "annotator",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Annotation"
    ws.append(headers)
    for row in rows:
        vals = []
        for h in headers:
            if h in {"gold_entity", "gold_location", "gold_time", "gold_event_type", "gold_relation", "gold_global", "gold_evidence_insufficient"}:
                vals.append("N")
                continue
            val = row.get(h, "")
            if isinstance(val, list):
                val = ", ".join(str(x) for x in val)
            vals.append(val)
        ws.append(vals)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in ws.columns:
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = 18
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["P"].width = 45
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    max_row = max(2, len(rows) + 1)
    dv_type = DataValidation(type="list", formula1='"' + ",".join(LABELS) + '"', allow_blank=True)
    dv_yn = DataValidation(type="list", formula1='"' + ",".join(YN) + '"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"' + ",".join(STATUS) + '"', allow_blank=True)
    dv_annotator = DataValidation(type="list", formula1='"' + ",".join(ANNOTATORS) + '"', allow_blank=True)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_yn)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_annotator)
    dv_type.add(f"H2:H{max_row}")
    for col in ["I", "J", "K", "L", "M", "N", "O"]:
        dv_yn.add(f"{col}2:{col}{max_row}")
    dv_status.add(f"Q2:Q{max_row}")
    dv_annotator.add(f"R2:R{max_row}")
    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    ap = argparse.ArgumentParser(description="Build human annotation candidates for real OOC attribution evaluation.")
    ap.add_argument("--context-pairs", default="outputs/cove_lite_context_pairs.jsonl")
    ap.add_argument("--predictions", default="outputs/field_nli_attribution_v2.jsonl")
    ap.add_argument("--output", default="examples/real_ooc_attribution_eval_candidates.jsonl")
    ap.add_argument("--xlsx", default="examples/real_ooc_attribution_eval_candidates.xlsx")
    ap.add_argument("--n", type=int, default=80)
    ap.add_argument("--seed", type=int, default=2026)
    ap.add_argument("--min-chars", type=int, default=25)
    args = ap.parse_args()

    rows = read_jsonl(Path(args.context_pairs))
    preds = {key_of(r): r for r in read_jsonl(Path(args.predictions))}
    candidates: List[Dict[str, Any]] = []
    for row in rows:
        if not is_ooc(row):
            continue
        cur = str(row.get("current_caption") or "").strip()
        tru = str(row.get("true_image_context") or "").strip()
        if len(cur) < args.min_chars or len(tru) < args.min_chars:
            continue
        k = key_of(row)
        pred = preds.get(k, {})
        out = {
            "sample_id": k,
            "image_id": row.get("image_id", ""),
            "text_id": row.get("text_id", ""),
            "domain": row.get("domain", ""),
            "split": row.get("split", ""),
            "current_caption": cur,
            "true_image_context": tru,
            "vdt_label": "OOC",
            "vdt_score": "",
            "pred_mismatch_type": pred.get("v2_mismatch_type", ""),
            "pred_conflict_fields": pred.get("v2_conflict_fields", []),
            "gold_mismatch_type": "",
            "gold_conflict_fields": [],
            "rationale": "",
            "annotation_status": "todo",
            "annotator": "",
            "candidate_score": round(info_score(row, pred), 4),
        }
        candidates.append(out)

    rng = random.Random(args.seed)
    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for c in candidates:
        groups[str(c.get("pred_mismatch_type") or "unknown")].append(c)
    for g in groups.values():
        g.sort(key=lambda x: (-float(x.get("candidate_score") or 0), str(x.get("sample_id"))))
    selected: List[Dict[str, Any]] = []
    while len(selected) < args.n and any(groups.values()):
        for key in sorted(groups):
            if groups[key] and len(selected) < args.n:
                selected.append(groups[key].pop(0))
    rng.shuffle(selected)
    write_jsonl(Path(args.output), selected)
    export_xlsx(Path(args.xlsx), selected)
    print(json.dumps({
        "output": args.output,
        "xlsx": args.xlsx,
        "records": len(selected),
        "pool": len(candidates),
        "pred_type_counts": dict(sorted({k: len(v) for k, v in groups.items()}.items())),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
